"""
EU Parliament Watch — recent-only multi-source monitor.

Runs on GitHub Actions. It checks recent official European Parliament sources,
searches keywords in HTML/PDF/XML/JSON content, exports matches to Excel, and
sends Telegram alerts with the Excel file attached.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import io
import json
import os
import re
import sys
import textwrap
import time
import urllib.parse
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

SEEN_FILE = DATA_DIR / "seen.json"
KEYWORDS_FILE = BASE_DIR / "keywords.txt"

USER_AGENT = (
    "eu-parliament-watch/1.0 "
    "(+https://github.com/; monitored personal research bot)"
)

REQUEST_TIMEOUT = 25
MAX_CANDIDATES = int(os.getenv("MAX_CANDIDATES", "6000"))
MAX_DOWNLOAD_BYTES = int(os.getenv("MAX_DOWNLOAD_BYTES", str(12 * 1024 * 1024)))
MAX_TEXT_CHARS = int(os.getenv("MAX_TEXT_CHARS", "300000"))
MAX_PDF_PAGES = int(os.getenv("MAX_PDF_PAGES", "80"))
MAX_ALERT_ITEMS = int(os.getenv("MAX_ALERT_ITEMS", "12"))
REQUEST_DELAY_SECONDS = float(os.getenv("REQUEST_DELAY_SECONDS", "0.15"))

# Current European Parliament committee codes to monitor explicitly.
# Includes committee-specific latest documents, meeting documents, votes, and work-in-progress pages.
COMMITTEE_CODES = [
    "AFET", "DROI", "SEDE", "DEVE", "INTA", "BUDG", "CONT", "ECON", "FISC",
    "EMPL", "ENVI", "SANT", "ITRE", "IMCO", "TRAN", "REGI", "AGRI", "PECH",
    "CULT", "JURI", "LIBE", "AFCO", "FEMM", "PETI", "EUDS", "HOUS",
]

# Committee document types explicitly covered. Codes are used defensively:
# if a URL stops returning useful results, the run logs it but continues.
COMMITTEE_DOCUMENT_TYPE_CODES = [
    "AM",   # Amendments
    "AB",   # Budget amendments
    "PR",   # Draft reports
    "RR",   # Reports
    "PA",   # Draft opinions
    "AD",   # Opinions
    "AG",   # Draft agendas / agendas
    "PV",   # Minutes
    "DT",   # Working documents
]

COMMITTEE_SEED_PAGES: list[str] = []
for _committee_code in COMMITTEE_CODES:
    _c = _committee_code.lower()
    COMMITTEE_SEED_PAGES.extend([
        f"https://www.europarl.europa.eu/committees/en/{_c}/documents/latest-documents",
        f"https://www.europarl.europa.eu/committees/en/{_c}/documents/search",
        f"https://www.europarl.europa.eu/committees/en/{_c}/documents/work-in-progress",
        f"https://www.europarl.europa.eu/committees/en/{_c}/meetings/meeting-documents",
        f"https://www.europarl.europa.eu/committees/en/{_c}/meetings/votes",
    ])
    # Recent-only search pages by document type, all committees.
    for _dtc in COMMITTEE_DOCUMENT_TYPE_CODES:
        COMMITTEE_SEED_PAGES.append(
            "https://www.europarl.europa.eu/committees/en/documents/search?"
            + urllib.parse.urlencode({
                "committeeMnemoCode": _committee_code,
                "documentTypeCode": _dtc,
                "performSearch": "true",
                "term": "10",
                "page": "0",
            })
        )

PLENARY_SEED_PAGES = [
    # Plenary current documents and votes
    "https://www.europarl.europa.eu/plenary/en/agendas.html",
    "https://www.europarl.europa.eu/plenary/en/documents.html",
    "https://www.europarl.europa.eu/plenary/en/votes.html",
    "https://www.europarl.europa.eu/plenary/en/votes.html?tab=votes",
    "https://www.europarl.europa.eu/plenary/en/minutes.html",
    "https://www.europarl.europa.eu/plenary/en/texts-adopted.html",
    "https://www.europarl.europa.eu/plenary/en/debates-video.html",
    "https://www.europarl.europa.eu/plenary/en/parliamentary-questions.html?tabType=wq",
]

OTHER_EP_SEED_PAGES = [
    # DOCEO / register / RSS / think tank / news: recent-entry pages only, no archive backfill.
    "https://www.europarl.europa.eu/doceo/recent-documents",
    "https://www.europarl.europa.eu/RegistreWeb/home/welcome.htm",
    "https://www.europarl.europa.eu/RegistreWeb/search/simpleSearchHome.htm",
    "https://www.europarl.europa.eu/at-your-service/en/stay-informed/rss-feeds",
    "https://www.europarl.europa.eu/news/en/press-room",
    "https://www.europarl.europa.eu/thinktank/en/home",
    "https://www.europarl.europa.eu/thinktank/en/research/advanced-search",
]

GLOBAL_COMMITTEE_PAGES = [
    "https://www.europarl.europa.eu/committees/en/documents/latest-documents",
    "https://www.europarl.europa.eu/committees/en/documents/search",
    "https://www.europarl.europa.eu/committees/en/meetings/meeting-documents",
]
for _dtc in COMMITTEE_DOCUMENT_TYPE_CODES:
    GLOBAL_COMMITTEE_PAGES.append(
        "https://www.europarl.europa.eu/committees/en/documents/search?"
        + urllib.parse.urlencode({
            "documentTypeCode": _dtc,
            "performSearch": "true",
            "term": "10",
            "page": "0",
        })
    )

# Recent-only: these are current/feed/latest pages, not archive backfill searches.
SEED_PAGES = []
SEED_PAGES.extend(PLENARY_SEED_PAGES)
SEED_PAGES.extend(GLOBAL_COMMITTEE_PAGES)
SEED_PAGES.extend(COMMITTEE_SEED_PAGES)
SEED_PAGES.extend(OTHER_EP_SEED_PAGES)

# Open Data API feed candidates. The code tries these defensively because formats
# and query parameters can change. Failed endpoints are logged but do not stop the run.
OPEN_DATA_FEEDS = [
    # Defensive list. Some may return 404 depending on EP API routing; failures are logged only.
    "https://data.europarl.europa.eu/api/v2/documents/feed",
    "https://data.europarl.europa.eu/api/v2/documents",
    "https://data.europarl.europa.eu/api/v2/plenary-documents/feed",
    "https://data.europarl.europa.eu/api/v2/plenary-documents",
    "https://data.europarl.europa.eu/api/v2/plenary-session-documents/feed",
    "https://data.europarl.europa.eu/api/v2/parliamentary-questions/feed",
    "https://data.europarl.europa.eu/api/v2/parliamentary-questions",
    "https://data.europarl.europa.eu/api/v2/committee-documents/feed",
    "https://data.europarl.europa.eu/api/v2/committee-documents",
]

INTERESTING_URL_PATTERNS = [
    "/doceo/document/",
    "/plenary/",
    "/committees/en/documents/",
    "/committees/en/meetings/",
    "/RegistreWeb/",
    "/thinktank/",
    "/news/",
    "data.europarl.europa.eu",
]

DOCUMENT_EXTENSIONS = (".pdf", ".xml", ".json", ".html", ".htm")


@dataclass
class Candidate:
    url: str
    source: str
    title: str = ""
    kind: str = "unknown"


@dataclass
class MatchResult:
    detected_at_utc: str
    source: str
    kind: str
    title: str
    url: str
    keywords: str
    excerpt: str
    content_hash: str
    status: str = "new"


def log(msg: str) -> None:
    print(f"[{dt.datetime.utcnow().isoformat(timespec='seconds')}Z] {msg}", flush=True)


def load_keywords() -> list[str]:
    if not KEYWORDS_FILE.exists():
        return ["Morocco", "Western Sahara", "Sahara occidental"]
    kws = []
    for line in KEYWORDS_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        kws.append(line)
    # longest first helps snippets around precise phrases
    return sorted(set(kws), key=len, reverse=True)


def load_seen() -> set[str]:
    if not SEEN_FILE.exists():
        return set()
    try:
        data = json.loads(SEEN_FILE.read_text(encoding="utf-8"))
        raw_items: list[str]
        if isinstance(data, list):
            raw_items = [str(x) for x in data]
        elif isinstance(data, dict):
            raw_items = [str(x) for x in data.get("seen", [])]
        else:
            raw_items = []
        seen = set(raw_items)
        # Backward compatibility with the first simple bot, which may have stored raw URLs.
        for item in raw_items:
            if item.startswith("http"):
                seen.add("url:" + url_key(item))
        return seen
    except Exception:
        return set()


def save_seen(seen: set[str]) -> None:
    payload = {
        "updated_at_utc": dt.datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "seen": sorted(seen),
    }
    SEEN_FILE.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*"})
    return s


def fetch(url: str, *, accept: str | None = None) -> tuple[bytes, str]:
    s = session()
    headers = {}
    if accept:
        headers["Accept"] = accept
    r = s.get(url, timeout=REQUEST_TIMEOUT, headers=headers, allow_redirects=True)
    r.raise_for_status()
    content_type = r.headers.get("content-type", "")
    content = r.content[:MAX_DOWNLOAD_BYTES]
    return content, content_type


def normalize_url(url: str, base: str | None = None) -> str | None:
    if not url:
        return None
    url = url.strip()
    if url.startswith("mailto:") or url.startswith("javascript:") or url.startswith("tel:"):
        return None
    if base:
        url = urllib.parse.urljoin(base, url)
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in ("http", "https"):
        return None
    # Remove fragments. Keep query parameters because PE pages often need them.
    parsed = parsed._replace(fragment="")
    return urllib.parse.urlunparse(parsed)


def is_ep_url(url: str) -> bool:
    host = urllib.parse.urlparse(url).netloc.lower()
    return host.endswith("europarl.europa.eu") or host.endswith("data.europarl.europa.eu")


def looks_interesting(url: str) -> bool:
    lower = url.lower()
    if not is_ep_url(url):
        return False
    if any(p.lower() in lower for p in INTERESTING_URL_PATTERNS):
        return True
    if lower.endswith(DOCUMENT_EXTENSIONS):
        return True
    return False


def infer_kind(url: str, title: str = "") -> str:
    text = f"{url} {title}".lower()
    if "roll-call" in text or "roll call" in text or "rcv" in text or "votes" in text or "vot" in text:
        return "roll-call vote / vote result"
    if "-asw" in text or "answer" in text:
        return "answer to written question"
    if "/doceo/document/e-" in text:
        return "written question"
    if "amend" in text or "/am-" in text:
        return "amendment"
    if "agenda" in text or "oj" in text:
        return "agenda"
    if "report" in text or "/rr-" in text or "a10-" in text:
        return "report"
    if "resolution" in text or "/ta-" in text:
        return "resolution / text adopted"
    if "committee" in text or "/committees/" in text:
        return "committee document"
    if "registreweb" in text:
        return "public register"
    if "plenary" in text:
        return "plenary document"
    if "thinktank" in text:
        return "think tank / research publication"
    if "/news/" in text or "press-room" in text:
        return "press / news publication"
    return "document"


def text_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()[:16]


def url_key(url: str) -> str:
    return hashlib.sha256(url.encode("utf-8", errors="ignore")).hexdigest()[:20]


def _contextual_link_title(a: Any) -> str:
    """Recover the document title around a PDF/DOC link.

    EP committee pages often label the anchor only as “PDF (137 KB)”. The real
    title is in the nearest preceding H3. Using only anchor text caused notices
    such as PETI_CM(2026)790891 to be missed even though “EU-Morocco” was in the
    visible title.
    """
    direct = (a.get_text(" ", strip=True) or a.get("title") or "").strip()
    generic = bool(re.fullmatch(r"(?i)(pdf|docx?|xml|html?)(\s*\([^)]*\))?", direct))
    if not generic and len(direct) > 20:
        return direct

    # Prefer the closest previous heading; this matches the current committee
    # “latest documents” cards where H3 precedes metadata and PDF/DOC links.
    heading = a.find_previous(["h1", "h2", "h3", "h4", "h5"])
    if heading:
        text = heading.get_text(" ", strip=True)
        if text:
            return text

    # Fallback to the closest card/list/article container.
    container = a.find_parent(["article", "li", "section", "div"])
    if container:
        text = re.sub(r"\s+", " ", container.get_text(" ", strip=True)).strip()
        if text:
            return text[:500]
    return direct


def extract_links_from_html(html: bytes, base_url: str) -> tuple[str, list[Candidate]]:
    soup = BeautifulSoup(html, "lxml")
    title = (soup.title.get_text(" ", strip=True) if soup.title else "").strip()
    links: list[Candidate] = []
    for a in soup.find_all("a", href=True):
        href = normalize_url(a.get("href"), base_url)
        if not href or not looks_interesting(href):
            continue
        link_title = _contextual_link_title(a)
        links.append(Candidate(url=href, source=base_url, title=link_title, kind=infer_kind(href, link_title)))
    return title, links


def html_to_text(html: bytes) -> str:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    return soup.get_text(" ", strip=True)[:MAX_TEXT_CHARS]


def pdf_to_text(content: bytes) -> str:
    try:
        reader = PdfReader(io.BytesIO(content))
        parts = []
        for page in reader.pages[:MAX_PDF_PAGES]:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n".join(parts)[:MAX_TEXT_CHARS]
    except Exception as exc:
        log(f"PDF extraction failed: {exc}")
        return ""


def xml_or_json_to_text(content: bytes) -> str:
    try:
        return content.decode("utf-8", errors="ignore")[:MAX_TEXT_CHARS]
    except Exception:
        return str(content[:MAX_TEXT_CHARS])


def keyword_hits(text: str, keywords: list[str]) -> list[str]:
    lower = text.lower()
    hits = []
    for kw in keywords:
        # phrase match, case-insensitive
        if kw.lower() in lower:
            hits.append(kw)
    return hits


def make_excerpt(text: str, hits: list[str], width: int = 550) -> str:
    if not text:
        return ""
    lower = text.lower()
    idx = -1
    for hit in hits:
        idx = lower.find(hit.lower())
        if idx >= 0:
            break
    if idx < 0:
        return text[:width].replace("\n", " ")
    start = max(0, idx - width // 3)
    end = min(len(text), idx + width)
    excerpt = text[start:end].replace("\n", " ")
    excerpt = re.sub(r"\s+", " ", excerpt).strip()
    if start > 0:
        excerpt = "…" + excerpt
    if end < len(text):
        excerpt += "…"
    return excerpt


def extract_urls_from_json(obj: Any) -> Iterable[str]:
    if isinstance(obj, dict):
        for key, val in obj.items():
            if isinstance(val, str):
                if val.startswith("http") or "/doceo/document/" in val:
                    yield val
            else:
                yield from extract_urls_from_json(val)
    elif isinstance(obj, list):
        for item in obj:
            yield from extract_urls_from_json(item)


def candidate_from_json_item(url: str, source: str) -> Candidate | None:
    nurl = normalize_url(url)
    if not nurl or not looks_interesting(nurl):
        return None
    return Candidate(url=nurl, source=source, title="Open Data feed item", kind=infer_kind(nurl))


def collect_from_open_data() -> list[Candidate]:
    candidates: list[Candidate] = []
    for base in OPEN_DATA_FEEDS:
        # Try multiple formats; some deployments accept one, not all.
        urls = [
            base,
            base + "?format=application/ld+json",
            base + "?format=json",
        ]
        for url in urls:
            try:
                content, ctype = fetch(url, accept="application/json, application/ld+json, text/turtle, */*")
                text = content.decode("utf-8", errors="ignore")
                # JSON-LD / JSON path
                if text.strip().startswith(("{", "[")):
                    data = json.loads(text)
                    for found in extract_urls_from_json(data):
                        cand = candidate_from_json_item(found, url)
                        if cand:
                            candidates.append(cand)
                else:
                    # Fallback: regex URLs from Turtle/RDF/plain text
                    for found in re.findall(r"https?://[^\s<>\"']+", text):
                        cand = candidate_from_json_item(found, url)
                        if cand:
                            candidates.append(cand)
                log(f"Open Data feed OK: {url}")
                break
            except Exception as exc:
                log(f"Open Data feed failed: {url}: {exc}")
                continue
    return candidates


def collect_recent_candidates(keywords: list[str]) -> list[Candidate]:
    all_candidates: list[Candidate] = []

    # 1) Open Data official feeds.
    all_candidates.extend(collect_from_open_data())

    # 2) Current/latest pages.
    for seed in SEED_PAGES:
        try:
            content, ctype = fetch(seed)
            if "html" not in ctype.lower() and not seed.lower().endswith((".html", ".htm")):
                continue
            title, links = extract_links_from_html(content, seed)
            # Search the seed page itself too, because some pages expose relevant text
            # without direct document links.
            all_candidates.append(Candidate(url=seed, source="seed page", title=title, kind=infer_kind(seed, title)))
            all_candidates.extend(links)
            log(f"Seed OK: {seed} -> {len(links)} links")
        except Exception as exc:
            log(f"Seed failed: {seed}: {exc}")

    # Deduplicate by normalized URL.
    dedup: dict[str, Candidate] = {}
    for c in all_candidates:
        nurl = normalize_url(c.url)
        if not nurl or not is_ep_url(nurl):
            continue
        if nurl not in dedup:
            c.url = nurl
            if not c.kind or c.kind == "unknown":
                c.kind = infer_kind(c.url, c.title)
            dedup[nurl] = c
    def priority(c: Candidate) -> tuple[int, int, str]:
        hay = f"{c.title} {c.url}".lower()
        title_hit = any(k.lower() in hay for k in keywords)
        direct_document = bool(re.search(r"(?i)(/RegData/|/doceo/document/|\.(pdf|xml|json|docx?)(?:$|\?))", c.url))
        # keyword-bearing titles first, then direct documents, then navigation pages
        return (0 if title_hit else 1, 0 if direct_document else 1, c.url)

    candidates = sorted(dedup.values(), key=priority)[:MAX_CANDIDATES]
    log(f"Collected {len(candidates)} unique recent candidates (prioritised)")
    return candidates


def fetch_candidate_text(candidate: Candidate) -> tuple[str, str, str]:
    """Return (text, title, content_hash)."""
    content, ctype = fetch(candidate.url)
    lower_url = candidate.url.lower()
    title = candidate.title or candidate.url

    if ".pdf" in lower_url or "pdf" in ctype.lower():
        text = pdf_to_text(content)
    elif ".xml" in lower_url or "xml" in ctype.lower():
        text = xml_or_json_to_text(content)
    elif ".json" in lower_url or "json" in ctype.lower():
        text = xml_or_json_to_text(content)
    else:
        text = html_to_text(content)
        try:
            html_title, _ = extract_links_from_html(content, candidate.url)
            if html_title and not candidate.title:
                title = html_title
        except Exception:
            pass
    return text, title, text_hash(candidate.url + "\n" + text[:10000])


def create_excel(results: list[MatchResult]) -> Path:
    date_str = dt.datetime.utcnow().strftime("%Y-%m-%d_%H%M")
    out = OUTPUT_DIR / f"ep_watch_results_{date_str}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"

    headers = [
        "Detected at UTC",
        "Source",
        "Type",
        "Title",
        "Keywords",
        "Excerpt",
        "URL",
        "Content hash",
        "Status",
    ]
    ws.append(headers)
    for r in results:
        ws.append([
            r.detected_at_utc,
            r.source,
            r.kind,
            r.title,
            r.keywords,
            r.excerpt,
            r.url,
            r.content_hash,
            r.status,
        ])

    # Second sheet specific to votes/RCV.
    vote_rows = [r for r in results if "vote" in r.kind.lower() or "roll-call" in r.kind.lower()]
    ws2 = wb.create_sheet("Votes_RCV_matches")
    ws2.append(headers)
    for r in vote_rows:
        ws2.append([
            r.detected_at_utc,
            r.source,
            r.kind,
            r.title,
            r.keywords,
            r.excerpt,
            r.url,
            r.content_hash,
            r.status,
        ])

    ws3 = wb.create_sheet("Readme")
    ws3.append(["Field", "Explanation"])
    ws3.append(["Detected at UTC", "Time when GitHub Actions found the match."])
    ws3.append(["Type", "Inferred document type from URL/title/content."])
    ws3.append(["Keywords", "Matched words/phrases from keywords.txt."])
    ws3.append(["Excerpt", "Text around the first matched keyword."])
    ws3.append(["Votes_RCV_matches", "Subset of matches inferred to be vote or roll-call vote documents. If the official XML structure changes, the raw source URL remains the reference."])

    # Formatting
    for sheet in [ws, ws2, ws3]:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = {
            "A": 20, "B": 28, "C": 24, "D": 48, "E": 30,
            "F": 70, "G": 70, "H": 18, "I": 12,
        }
        for col, width in widths.items():
            sheet.column_dimensions[col].width = width
        for row_idx in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row_idx].height = 45
        sheet.auto_filter.ref = sheet.dimensions

    wb.save(out)
    return out


def send_telegram_message(text: str) -> None:
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        log("Telegram secrets not set; skipping message")
        return
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = {
        "chat_id": chat_id,
        "text": text[:3900],
        "disable_web_page_preview": "false",
    }
    r = requests.post(url, data=data, timeout=REQUEST_TIMEOUT)
    if not r.ok:
        log(f"Telegram sendMessage failed: {r.status_code} {r.text[:200]}")


def send_telegram_document(path: Path, caption: str) -> None:
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        log("Telegram secrets not set; skipping document")
        return
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with path.open("rb") as f:
        files = {"document": (path.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"chat_id": chat_id, "caption": caption[:1000]}
        r = requests.post(url, data=data, files=files, timeout=REQUEST_TIMEOUT)
    if not r.ok:
        log(f"Telegram sendDocument failed: {r.status_code} {r.text[:200]}")


def build_simple_alert(result: MatchResult, excel_path: Path) -> str:
    """Return the compact Telegram alert format preferred by the user."""
    excerpt = result.excerpt.strip()
    if len(excerpt) > 1200:
        excerpt = excerpt[:1197] + "…"
    return (
        "🚨 Alerte Parlement européen\n\n"
        f"Type: {result.kind}\n"
        f"Mots-clés: {result.keywords}\n"
        f"Lien: {result.url}\n\n"
        f"Extrait: {excerpt}\n\n"
        f"📊 Export Excel généré: {excel_path}"
    )


def main() -> int:
    keywords = load_keywords()
    if not keywords:
        log("No keywords found. Add at least one keyword to keywords.txt")
        return 1

    seen = load_seen()
    new_seen = set(seen)
    candidates = collect_recent_candidates(keywords)
    results: list[MatchResult] = []
    detected_at = dt.datetime.utcnow().isoformat(timespec="seconds") + "Z"

    for idx, cand in enumerate(candidates, start=1):
        # Deduplicate URL before fetching. If URL has been seen and no content hash is
        # available, skip to save time. Content hashes are added when a match happens.
        ukey = "url:" + url_key(cand.url)
        if ukey in seen:
            continue
        title_combined = "\n".join([cand.title or "", cand.url, cand.kind or ""])
        title_hits = keyword_hits(title_combined, keywords)
        try:
            if idx % 50 == 0:
                log(f"Scanning candidate {idx}/{len(candidates)}")
            text, title, chash = fetch_candidate_text(cand)
            combined = "\n".join([title or cand.title or "", cand.url, cand.kind or "", text or ""])
            hits = keyword_hits(combined, keywords)
            new_seen.add(ukey)
            if not hits:
                continue
            ckey = "content:" + chash
            if ckey in seen:
                continue
            new_seen.add(ckey)
            result = MatchResult(
                detected_at_utc=detected_at,
                source=cand.source,
                kind=infer_kind(cand.url, title or cand.title),
                title=title or cand.title or cand.url,
                url=cand.url,
                keywords=", ".join(hits),
                excerpt=make_excerpt(combined, hits),
                content_hash=chash,
            )
            results.append(result)
            log(f"MATCH: {result.kind} {result.keywords} {result.url}")
            time.sleep(REQUEST_DELAY_SECONDS)
        except Exception as exc:
            # Never mark an inaccessible document as seen: a temporary 403/timeout
            # must be retried at the next run. If the title itself already matches,
            # alert immediately because title-level detection is sufficient.
            log(f"Candidate failed: {cand.url}: {exc}")
            if title_hits:
                chash = text_hash(cand.url + "\n" + (cand.title or ""))
                ckey = "content:" + chash
                if ckey not in seen:
                    new_seen.add(ukey)
                    new_seen.add(ckey)
                    results.append(MatchResult(
                        detected_at_utc=detected_at,
                        source=cand.source,
                        kind=infer_kind(cand.url, cand.title),
                        title=cand.title or cand.url,
                        url=cand.url,
                        keywords=", ".join(title_hits),
                        excerpt=make_excerpt(title_combined, title_hits),
                        content_hash=chash,
                        status="title match; content fetch failed",
                    ))
            continue

    if results:
        excel_path = create_excel(results)
        for result in results[:MAX_ALERT_ITEMS]:
            send_telegram_message(build_simple_alert(result, excel_path))
        if len(results) > MAX_ALERT_ITEMS:
            send_telegram_message(
                f"🚨 Alerte Parlement européen\n\n"
                f"{len(results) - MAX_ALERT_ITEMS} autre(s) résultat(s) dans l'Excel.\n"
                f"📊 Export Excel généré: {excel_path}"
            )
        send_telegram_document(excel_path, f"📊 Export Excel généré: {excel_path.name}")
        log(f"Export Excel generated: {excel_path}")
    else:
        log("No new keyword matches found.")

    save_seen(new_seen)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
